
from selenium import webdriver
import time
from selenium.webdriver.common.by import By
import openpyxl
from datetime import datetime as date


# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    # To run Chrome browser. If you get error for not finding the browser, set driver = webdriver.Chrome(".\Browsers\chromedriver.exe").
    driver = webdriver.Chrome()

    # To maximize the browser window
    driver.maximize_window()

    # To search www.google.com in the chrome
    driver.get("https://www.google.com/")

    # To load the Excel file
    data = openpyxl.load_workbook(".\Data\Excel.xlsx")

    # To get today's day
    today_date = date.today().strftime("%A")

    # To select today's sheet
    read = data[today_date]

    # Search the Keyword
    for i in range(10):
        # Get the keyword from sheet
        cell_value = read.cell(row=3 + i, column=3).value

        # Send the keyword in search dialog
        driver.find_element(By.NAME, "q").send_keys(cell_value)
        time.sleep(3)

        # Get the auto-suggestion list
        suggestion_list = driver.find_elements(By.XPATH, "//div[@class='pcTkSc']//div[1]//span")

        # Find the largest and smallest suggestion
        max_size, min_size = 0, 100
        max_str, min_str = " ", " "
        for suggestions in suggestion_list:
            if max_size < len(suggestions.text):
                max_size = len(suggestions.text)
                max_str = suggestions.text
            if min_size > len(suggestions.text) > 0:
                min_size = len(suggestions.text)
                min_str = suggestions.text

        # Store the result in the sheet
        read.cell(row=3 + i, column=4).value = max_str
        read.cell(row=3 + i, column=5).value = min_str

        # Clear the search dialog
        driver.find_element(By.NAME, "q").clear()

    # Save the sheet in the Excel file
    data.save(".\Data\Excel.xlsx")

    # Close the driver
    driver.close()

# See PyCharm help at https://www.jetbrains.com/help/pycharm/
